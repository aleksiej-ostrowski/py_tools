#-----------------------------------------------------#
#                                                     #
#  The simplest way to export list of lists to Excel  #
#                                                     #
#  version 0.0.1                                      #
#                                                     #
#  Aleksiej Ostrowski, 2021                           #
#                                                     #
#  https://aleksiej.com                               #
#                                                     #
#-----------------------------------------------------#

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

import random
import time

class ExcelColorTable:
    def __init__(self, table, title, filename, fast = False):

        self.table = table
        self.filename = filename

        self.workbook = Workbook()
        sheet = self.workbook.active

        sheet.title = title
        # sheet.sheet_properties.tabColor = "1072BA"

        for row in table:
            sheet.append(row)

        sheet.freeze_panes = "A2"

        if (not fast):

            col_range = sheet.max_column
            row_range = sheet.max_row

            for rows in sheet.iter_rows(min_row = 1, max_row = 1, min_col = 1):
                for i, cell in enumerate(rows):

                    sheet.column_dimensions[get_column_letter(i + 1)].width = 20.0

                    fill_color = mutate_colors('63D378', float(i) / col_range)  
                    font_color = '004800'

                    if i % 2 == 0:
                        fill_color = mutate_colors('FFD009', float(i) / col_range)  
                        font_color = '825400'

                    cell.fill = PatternFill(start_color = fill_color, end_color = fill_color, patternType = 'solid')
                    cell.font = Font(color = font_color, size = 12)

            for rows in sheet.iter_rows(min_row = 2, max_row = row_range, min_col = 1):
                for i, cell in enumerate(rows):

                    fill_color = mutate_colors('C6EFCE', float(i) / col_range)
                    font_color = '006100'

                    if i % 2 == 0:
                        fill_color = mutate_colors('FFEB9C', float(i) / col_range)
                        font_color = '9C6500'

                    cell.fill = PatternFill(start_color = fill_color, end_color = fill_color, patternType = 'solid')
                    cell.font = Font(color = font_color, size = 12)

    def __del__(self):
        self.workbook.close()

    def savetable(self):
        self.workbook.save(self.filename) 

def mutate_colors(color, s):
    data = list(color)
    data = [str(i) + str(k) for i, k in zip (data[0::2], data[1::2])]
    random.seed(s)
    random.shuffle(data)
    return ''.join(data)
        
def main():
    random.seed(time.time())
    table = [[random.randint(1, 99) for e in range(100)] for x in range(1000)]

    ECT = ExcelColorTable(table = table, title = "example", filename = "example.xls")
    ECT.savetable()

if __name__ == '__main__':
    main()
